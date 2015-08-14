#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os, sys
import xlrd
import openpyxl
from logutils import logError,logINFO

'''
Excel解析处理公用方法
'''

#打开excel文件
def open_excel(path):
	data = None
	try:
		data = xlrd.open_workbook(path)
	except Exception, e:
		logError('获取excel数据Error', e)
	return data
		
#根据sheet名称获取Excel中的数据
def data_of_excel(path, sheet, cols):
	data = open_excel(path)
	table = data.sheet_by_name(sheet)
	#行数,列
	nrows = table.nrows
	nclus = cols
	list = []
	for row in range(0, nrows):
		list.append(table.row_values(row))
	return list
	
'''
解析Excel (.xlsx)
@param path 	文件路径
@param sheet 	sheetnames
'''
def xlsxParse(path, sheet):
	sheetNum = 0
	data_dic = {}
	# Load File
	wb = openpyxl.load_workbook(path)
	# 获取sheet
	sheetnames = wb.get_sheet_names()
	for index,name in enumerate(sheetnames):
		if name == sheet:
			sheetNum = index
			break
	print '................'
	try:
		ws = wb.get_sheet_by_name(sheetnames[sheetNum])
		rownum = ws.get_highest_row() + 1
		cellnum = ws.get_highest_column() + 1

		print rownum, cellnum
		for r in xrange(1, rownum):
			temp_list = []
			for c in xrange(1, cellnum):
				val = ws.cell(row=r,column=c).value
				if val != None:
					temp_list.append(val)
			data_dic[r] = temp_list
	except Exception, e:
		logError("解析EXCEL报错", e)
	return data_dic

if __name__ == '__main__':
	print xlsxParse('E:/dumps/xh_20150813.xlsx', u'苹果GS')